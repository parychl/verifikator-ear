from datetime import datetime, timezone
from pathlib import Path

from ear_verifikator.core import export
from ear_verifikator.core.model import (
    Chyba,
    InfoEAR,
    InfoPDFA,
    InfoPodpisu,
    InfoRazitka,
    KodChyby,
    Vysledek,
    VysledekSouboru,
)


def _vysledek_platny() -> VysledekSouboru:
    return VysledekSouboru(
        soubor=Path("C:/dokumentace/vykres1.pdf"),
        vysledek=Vysledek.PLATNY,
        pdfa=InfoPDFA(deklarovana="PDF/A-3B", vyhovuje=True),
        podpisy=[
            InfoPodpisu(
                pole="Signature1",
                podepsal="Ing. Jan Novák",
                vydavatel="PostSignum Qualified CA 4",
                cas_podpisu=datetime(2026, 4, 22, 9, 30, tzinfo=timezone.utc),
                sub_filter="/ETSI.CAdES.detached",
                integrita=True,
                pokryva_dokument=True,
                duveryhodny_retezec=True,
                kvalifikovany_cert=True,
                ear=InfoEAR(True, "Ing. Jan Novák", "0012345", "pozemní stavby", "ČKAIT"),
                razitko=InfoRazitka(
                    pritomno=True,
                    cas=datetime(2026, 4, 22, 9, 31, tzinfo=timezone.utc),
                    tsa="I.CA TSA",
                    podpis_platny=True,
                    kvalifikovane=True,
                ),
            )
        ],
    )


def _vysledek_neplatny() -> VysledekSouboru:
    return VysledekSouboru(
        soubor=Path("C:/dokumentace/vykres2.pdf"),
        vysledek=Vysledek.NEPLATNY,
        chyby=[Chyba(KodChyby.CHYBI_CASOVE_RAZITKO), Chyba(KodChyby.PODPIS_NENI_EAR)],
        pdfa=InfoPDFA(deklarovana="", vyhovuje=False, detail="nedeklaruje PDF/A"),
    )


VYSLEDKY = [_vysledek_platny(), _vysledek_neplatny()]


def test_filtruj():
    assert len(export.filtruj(VYSLEDKY, "vse")) == 2
    assert [v.vysledek for v in export.filtruj(VYSLEDKY, "platne")] == [Vysledek.PLATNY]
    assert [v.vysledek for v in export.filtruj(VYSLEDKY, "neplatne")] == [Vysledek.NEPLATNY]


def test_radek_obsahuje_vse():
    radek = export.radek_souboru(_vysledek_platny())
    assert set(radek) == set(export.SLOUPCE)
    assert radek["Číslo autorizace"] == "0012345"
    assert radek["Komora"] == "ČKAIT"
    assert radek["Razítko kvalifikované"] == "ano"
    assert radek["PDF/A splňuje"] == "ano"


def test_export_csv(tmp_path):
    cesta = tmp_path / "report.csv"
    export.exportuj(cesta, VYSLEDKY)
    obsah = cesta.read_text(encoding="utf-8-sig")
    assert obsah.startswith("Soubor;")
    assert "Chybí časové razítko" in obsah
    assert "Objednat kvalifikované časové razítko" in obsah  # doporučení
    assert obsah.count("vykres") == 4  # 2 soubory × (název + cesta)


def test_export_xlsx(tmp_path):
    from openpyxl import load_workbook

    cesta = tmp_path / "report.xlsx"
    export.exportuj(cesta, VYSLEDKY)
    ws = load_workbook(cesta).active
    assert ws.max_row == 3  # hlavička + 2 soubory
    assert ws.cell(row=1, column=1).value == "Soubor"
    assert ws.cell(row=2, column=3).value == "Platný"
    assert "Chybí časové razítko" in ws.cell(
        row=3, column=export.SLOUPCE.index("Chyby a doporučení") + 1
    ).value


def test_export_txt(tmp_path):
    cesta = tmp_path / "report.txt"
    export.exportuj(cesta, VYSLEDKY)
    obsah = cesta.read_text(encoding="utf-8")
    assert "VÝSLEDEK: Platný" in obsah
    assert "VÝSLEDEK: Neplatný" in obsah
    assert "Doporučení:" in obsah


def test_export_bez_vzorcu(tmp_path):
    """Hodnoty z certifikátu (CN…) nesmí Excel vyhodnotit jako vzorec."""
    from openpyxl import load_workbook

    v = _vysledek_platny()
    v.podpisy[0].podepsal = '=HYPERLINK("http://utocnik.example","klik")'

    cesta_csv = tmp_path / "inj.csv"
    export.exportuj(cesta_csv, [v])
    obsah = cesta_csv.read_text(encoding="utf-8-sig")
    assert "'=HYPERLINK" in obsah  # neutralizováno apostrofem
    assert ";=HYPERLINK" not in obsah  # hodnota nikde nezačíná rovnítkem

    cesta_xlsx = tmp_path / "inj.xlsx"
    export.exportuj(cesta_xlsx, [v])
    ws = load_workbook(cesta_xlsx).active
    bunka = ws.cell(row=2, column=export.SLOUPCE.index("Podepsal") + 1)
    assert bunka.data_type != "f"  # uloženo jako text, ne jako vzorec
    assert bunka.value.startswith("=HYPERLINK")


def test_nepodporovany_format(tmp_path):
    import pytest

    with pytest.raises(ValueError):
        export.exportuj(tmp_path / "report.pdf", VYSLEDKY)
