"""Export výsledků kontroly do XLSX / CSV / TXT.

Export obsahuje všechny kontrolované údaje včetně plného popisu chyb
a doporučení k nápravě, aby šel report poslat zpracovateli dokumentace.
"""
from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from .model import Vysledek, VysledekSouboru

SLOUPCE = [
    "Soubor",
    "Cesta",
    "Výsledek",
    "Chyby a doporučení",
    "Podepsal",
    "EAR",
    "Komora",
    "Číslo autorizace",
    "Obor",
    "Kvalifikovaný certifikát",
    "Vydavatel certifikátu",
    "Čas podpisu",
    "Časové razítko",
    "Razítko kvalifikované",
    "Integrita podpisu",
    "Formát PAdES",
    "Uzamčen",
    "PDF/A deklarováno",
    "PDF/A splňuje",
    "PDF/A detail",
    "Podpisová pole",
]

# Buňky začínající těmito znaky Excel/LibreOffice vyhodnocují jako vzorce.
# Hodnoty pocházejí i z certifikátů a názvů souborů (ovlivnitelné autorem
# kontrolovaného PDF), proto se nesmí dostat do tabulky jako vzorec.
_ZACATKY_VZORCE = ("=", "+", "-", "@", "\t", "\r")


def _bez_vzorce_csv(hodnota: str) -> str:
    if hodnota.startswith(_ZACATKY_VZORCE):
        return "'" + hodnota
    return hodnota


def _ano_ne(hodnota: bool) -> str:
    return "ano" if hodnota else "ne"


def _fmt_cas(dt) -> str:
    if not dt:
        return ""
    if dt.tzinfo is not None:
        dt = dt.astimezone()  # jednotně v místním čase (razítka bývají v UTC)
    return dt.strftime("%d.%m.%Y %H:%M:%S")


def _chyby_text(v: VysledekSouboru) -> str:
    radky = []
    for ch in v.chyby:
        radek = ch.nazev
        if ch.detail:
            radek += f" — {ch.detail}"
        if ch.doporuceni:
            radek += f" | Doporučení: {ch.doporuceni}"
        radky.append(radek)
    if v.poznamka:
        radky.append(v.poznamka)
    return "\n".join(radky)


def _spoj(hodnoty) -> str:
    """Spojí neprázdné unikátní hodnoty přes '; ' se zachováním pořadí."""
    videne: list[str] = []
    for h in hodnoty:
        if h and h not in videne:
            videne.append(h)
    return "; ".join(videne)


def radek_souboru(v: VysledekSouboru) -> dict[str, str]:
    podpisy = [p for p in v.podpisy if not p.je_docasove_razitko]
    doc_ts = [p for p in v.podpisy if p.je_docasove_razitko]
    razitka = [p.razitko for p in v.podpisy if p.razitko.pritomno]

    if v.pdfa.vyhovuje is True:
        pdfa_splnuje = "ano"
    elif v.pdfa.vyhovuje is False:
        pdfa_splnuje = "ne"
    else:
        pdfa_splnuje = "neověřeno"

    return {
        "Soubor": v.soubor.name,
        "Cesta": str(v.soubor),
        "Výsledek": v.vysledek.value,
        "Chyby a doporučení": _chyby_text(v),
        "Podepsal": _spoj(p.podepsal for p in podpisy),
        "EAR": _spoj(_ano_ne(p.ear.je_ear) for p in podpisy),
        "Komora": _spoj(p.ear.komora for p in podpisy),
        "Číslo autorizace": _spoj(p.ear.cislo_autorizace for p in podpisy),
        "Obor": _spoj(p.ear.obor for p in podpisy),
        "Kvalifikovaný certifikát": _spoj(_ano_ne(p.kvalifikovany_cert) for p in podpisy),
        "Vydavatel certifikátu": _spoj(p.vydavatel for p in podpisy),
        "Čas podpisu": _spoj(_fmt_cas(p.cas_podpisu) for p in podpisy),
        "Časové razítko": _spoj(
            f"{_fmt_cas(r.cas)} ({r.tsa})".strip() for r in razitka
        ),
        "Razítko kvalifikované": _spoj(_ano_ne(r.kvalifikovane) for r in razitka),
        "Integrita podpisu": _spoj(
            "neporušen" if p.integrita and p.pokryva_dokument else "porušen"
            for p in podpisy
        ),
        "Formát PAdES": _spoj(p.sub_filter for p in podpisy),
        "Uzamčen": _ano_ne(v.uzamcen),
        "PDF/A deklarováno": v.pdfa.deklarovana or "nedeklaruje",
        "PDF/A splňuje": pdfa_splnuje,
        "PDF/A detail": v.pdfa.detail,
        "Podpisová pole": _spoj(
            [p.pole for p in podpisy]
            + [f"{p.pole} (dokumentové razítko)" for p in doc_ts]
        ),
    }


def filtruj(
    vysledky: list[VysledekSouboru], rozsah: str
) -> list[VysledekSouboru]:
    """rozsah: 'vse' | 'platne' | 'neplatne' (neplatné = vše kromě Platný)."""
    if rozsah == "platne":
        return [v for v in vysledky if v.vysledek == Vysledek.PLATNY]
    if rozsah == "neplatne":
        return [v for v in vysledky if v.vysledek != Vysledek.PLATNY]
    return list(vysledky)


def export_csv(cesta: Path, vysledky: list[VysledekSouboru]) -> None:
    # utf-8-sig + středník: výchozí formát pro české Excel prostředí
    with cesta.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=SLOUPCE, delimiter=";")
        w.writeheader()
        for v in vysledky:
            w.writerow(
                {k: _bez_vzorce_csv(h) for k, h in radek_souboru(v).items()}
            )


def export_xlsx(cesta: Path, vysledky: list[VysledekSouboru]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Kontrola EAR"
    ws.append(SLOUPCE)
    for bunka in ws[1]:
        bunka.font = Font(bold=True)
        bunka.alignment = Alignment(vertical="center")

    BARVY_VYSLEDKU = {
        Vysledek.PLATNY.value: "C8E6C9",
        Vysledek.NEPLATNY.value: "FFCDD2",
        Vysledek.TECHNICKY_NEVYHOVUJICI.value: "FFE0B2",
        Vysledek.ZASTARALY_STANDARD.value: "FFECB3",
        Vysledek.CHYBA_ZPRACOVANI.value: "E0E0E0",
    }

    for v in vysledky:
        radek = radek_souboru(v)
        ws.append([radek[s] for s in SLOUPCE])
        # openpyxl ukládá řetězce začínající '=' jako vzorce — vynutit text
        for bunka in ws[ws.max_row]:
            if bunka.data_type == "f":
                bunka.data_type = "s"
        vysledek_bunka = ws.cell(row=ws.max_row, column=SLOUPCE.index("Výsledek") + 1)
        barva = BARVY_VYSLEDKU.get(radek["Výsledek"])
        if barva:
            vysledek_bunka.fill = PatternFill("solid", fgColor=barva)
        chyby_bunka = ws.cell(
            row=ws.max_row, column=SLOUPCE.index("Chyby a doporučení") + 1
        )
        chyby_bunka.alignment = Alignment(wrap_text=True, vertical="top")

    sirky = {"Soubor": 32, "Cesta": 45, "Výsledek": 20, "Chyby a doporučení": 70,
             "PDF/A detail": 40, "Vydavatel certifikátu": 26, "Časové razítko": 34}
    for i, nazev in enumerate(SLOUPCE, start=1):
        ws.column_dimensions[get_column_letter(i)].width = sirky.get(nazev, 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(cesta)


def export_txt(cesta: Path, vysledky: list[VysledekSouboru]) -> None:
    """Čitelný textový report — vhodný k vložení do e-mailu zpracovateli."""
    radky = [
        "KONTROLA ELEKTRONICKÝCH AUTORIZAČNÍCH RAZÍTEK (EAR)",
        f"Vytvořeno: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        f"Souborů: {len(vysledky)}",
        "=" * 78,
    ]
    for v in vysledky:
        radek = radek_souboru(v)
        radky += ["", f"SOUBOR: {radek['Soubor']}", f"Cesta: {radek['Cesta']}",
                  f"VÝSLEDEK: {radek['Výsledek']}"]
        if radek["Chyby a doporučení"]:
            radky.append("Chyby:")
            radky += [f"  - {ch}" for ch in radek["Chyby a doporučení"].split("\n")]
        for nazev in SLOUPCE[4:]:
            if radek[nazev]:
                radky.append(f"{nazev}: {radek[nazev]}")
        radky.append("-" * 78)
    cesta.write_text("\n".join(radky) + "\n", encoding="utf-8")


EXPORTERY = {
    ".xlsx": export_xlsx,
    ".csv": export_csv,
    ".txt": export_txt,
}


def exportuj(cesta: Path, vysledky: list[VysledekSouboru]) -> None:
    pripona = cesta.suffix.lower()
    exporter = EXPORTERY.get(pripona)
    if exporter is None:
        raise ValueError(f"Nepodporovaný formát: {pripona} (podporováno: xlsx, csv, txt)")
    exporter(cesta, vysledky)
